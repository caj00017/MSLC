"""Import / export helpers: CSV, JSON, and a one-click Excel workbook.

CSV and JSON use only the standard library, so scenario round-tripping works
with a bare interpreter. The Excel export uses ``openpyxl`` if it is installed
and otherwise raises a clear, actionable error (the rest of the tool is
unaffected).

A "scenario file" is a single JSON bundle of assumptions + asset catalog +
dependency table, so a whole study can be shared as one artifact.
"""

from __future__ import annotations

import csv
import io
import json
import os
from typing import Any, Dict, List

from . import defaults


# --- typing helpers ---------------------------------------------------------

def _coerce(value: str, numeric: bool, boolean: bool) -> Any:
    if boolean:
        return str(value).strip().lower() in ("1", "true", "yes", "y", "t")
    if numeric:
        s = str(value).strip()
        if s == "" or s.lower() in ("nan", "none"):
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return value


# --- assumptions (flat dict) <-> JSON --------------------------------------

def assumptions_to_json(assumptions: Dict[str, Any], indent: int = 2) -> str:
    return json.dumps(assumptions, indent=indent, sort_keys=True)


def assumptions_from_json(text: str) -> Dict[str, Any]:
    data = json.loads(text)
    if not isinstance(data, dict):
        raise ValueError("Assumptions JSON must be a flat object.")
    merged = dict(defaults.DEFAULT_ASSUMPTIONS)
    merged.update(data)
    return merged


# --- asset catalog <-> CSV --------------------------------------------------

def assets_to_csv(assets: List[Dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=defaults.ASSET_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in assets:
        writer.writerow({col: row.get(col, "") for col in defaults.ASSET_COLUMNS})
    return buf.getvalue()


def assets_from_csv(text: str) -> List[Dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for raw in reader:
        row: Dict[str, Any] = {}
        for col in defaults.ASSET_COLUMNS:
            row[col] = _coerce(
                raw.get(col, ""),
                numeric=col in defaults.ASSET_NUMERIC_COLUMNS,
                boolean=col in defaults.ASSET_BOOL_COLUMNS,
            )
        rows.append(row)
    return rows


# --- dependency table <-> CSV ----------------------------------------------

def dependencies_to_csv(deps: List[Dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=defaults.DEPENDENCY_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in deps:
        writer.writerow({col: row.get(col, "") for col in defaults.DEPENDENCY_COLUMNS})
    return buf.getvalue()


def dependencies_from_csv(text: str) -> List[Dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for raw in reader:
        rows.append({
            "asset_id": raw.get("asset_id", "").strip(),
            "depends_on_asset_id": raw.get("depends_on_asset_id", "").strip(),
            "rule": int(_coerce(raw.get("rule", "0"), numeric=True, boolean=False)),
            "note": raw.get("note", ""),
        })
    return rows


# --- whole-scenario bundle <-> JSON ----------------------------------------

def scenario_to_json(assumptions: Dict[str, Any], assets: List[Dict[str, Any]],
                     dependencies: List[Dict[str, Any]], indent: int = 2) -> str:
    return json.dumps({
        "format": "mars_logistics_scenario",
        "version": 1,
        "assumptions": assumptions,
        "assets": assets,
        "dependencies": dependencies,
    }, indent=indent)


def scenario_from_json(text: str):
    data = json.loads(text)
    assumptions = dict(defaults.DEFAULT_ASSUMPTIONS)
    assumptions.update(data.get("assumptions", {}))
    assets = data.get("assets") or defaults.default_assets()
    dependencies = data.get("dependencies") or defaults.default_dependencies()
    return assumptions, assets, dependencies


# --- results / manifest -> CSV ---------------------------------------------

def dashboard_to_csv(results: Dict[str, Any]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["metric", "value"])
    for key, value in results["dashboard"].items():
        if key == "top_five_mass_drivers":
            value = "; ".join(f"{b}={m:.0f}" for b, m in value)
        writer.writerow([key, value])
    return buf.getvalue()


def manifest_to_csv(results: Dict[str, Any]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["mission", "item_id", "asset_name", "quantity", "mass_kg",
                     "volume_m3", "bucket", "criticality", "pre_crew", "partial"])
    for m in results["missions"]["manifest"]:
        for it in m.items:
            writer.writerow([m.index + 1, it.item_id, it.asset_name,
                             round(it.quantity, 3), round(it.mass_kg, 2),
                             round(it.volume_m3, 3), it.bucket, it.criticality,
                             it.must_arrive_before_crew, it.is_partial])
    return buf.getvalue()


def mass_by_category_to_csv(results: Dict[str, Any]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["category", "mass_kg"])
    for bucket, mass in results["mass"]["buckets"].items():
        writer.writerow([bucket, round(mass, 2)])
    return buf.getvalue()


# --- Excel workbook (optional dependency) ----------------------------------

def export_excel(results: Dict[str, Any], assets: List[Dict[str, Any]] = None) -> bytes:
    """Build a multi-sheet Excel workbook of the key outputs. Requires openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError(
            "Excel export needs openpyxl. Install it with `pip install openpyxl`, "
            "or use the CSV/JSON exports which need no extra packages."
        ) from exc

    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["metric", "value"])
    for key, value in results["dashboard"].items():
        if key == "top_five_mass_drivers":
            value = "; ".join(f"{b}={m:.0f}" for b, m in value)
        ws.append([key, value])

    ws = wb.create_sheet("Mass by category")
    ws.append(["category", "mass_kg"])
    for bucket, mass in results["mass"]["buckets"].items():
        ws.append([bucket, round(mass, 2)])

    ws = wb.create_sheet("Mission manifest")
    ws.append(["mission", "item_id", "asset_name", "quantity", "mass_kg",
               "volume_m3", "bucket", "criticality", "pre_crew", "partial"])
    for m in results["missions"]["manifest"]:
        for it in m.items:
            ws.append([m.index + 1, it.item_id, it.asset_name, round(it.quantity, 3),
                       round(it.mass_kg, 2), round(it.volume_m3, 3), it.bucket,
                       it.criticality, it.must_arrive_before_crew, it.is_partial])

    ws = wb.create_sheet("Power balance")
    p = results["power"]
    for k in ("assets_continuous_kw", "habitat_continuous_kw", "eclss_power_kw",
              "isru_power_kw", "thermal_power_kw", "total_continuous_kw", "total_peak_kw",
              "total_energy_kwh_per_sol", "required_generation_kw",
              "installed_continuous_kw", "power_margin_kw", "critical_load_kw"):
        ws.append([k, round(p.get(k, 0), 3)])

    ws = wb.create_sheet("Readiness")
    ws.append(["check", "category", "status", "value", "requirement"])
    for c in results["readiness"]["crew_checks"] + results["readiness"]["sustainment_checks"]:
        ws.append([c["name"], c["category"], c["status"], c["value"], c["requirement"]])

    if assets is not None:
        ws = wb.create_sheet("Asset catalog")
        ws.append(defaults.ASSET_COLUMNS)
        for row in assets:
            ws.append([row.get(col, "") for col in defaults.ASSET_COLUMNS])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --- default data-file generation ------------------------------------------

def write_default_data_files(data_dir: str) -> List[str]:
    """Generate data/ files from the canonical Python defaults. Returns paths."""
    os.makedirs(data_dir, exist_ok=True)
    written: List[str] = []

    paths = {
        "default_assumptions.json": assumptions_to_json(defaults.default_assumptions()),
        "default_assets.csv": assets_to_csv(defaults.default_assets()),
        "default_dependencies.csv": dependencies_to_csv(defaults.default_dependencies()),
    }
    for name, content in paths.items():
        path = os.path.join(data_dir, name)
        with open(path, "w", encoding="utf-8", newline="") as fh:
            fh.write(content)
        written.append(path)

    # The default mission manifest is the legacy reference manifest (clearly
    # labeled), kept as the seed manifest CSV the spec asks for.
    from .legacy import LEGACY_MISSION_MANIFEST
    manifest_path = os.path.join(data_dir, "default_mission_manifest.csv")
    with open(manifest_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["mission", "year", "lander", "mode", "launch_mass_limit_kg",
                         "allocated_mass_kg", "available_power_kw", "power_needed_kw",
                         "available_energy_kwh", "energy_needed_kwh", "key_assets",
                         "source_class"])
        for m in LEGACY_MISSION_MANIFEST:
            writer.writerow([m.mission, m.year, m.lander, m.mode, m.launch_mass_limit_kg,
                             m.allocated_mass_kg, m.available_power_kw, m.power_needed_kw,
                             m.available_energy_kwh, m.energy_needed_kwh, m.key_assets,
                             "legacy_lunar_analog_placeholder"])
    written.append(manifest_path)
    return written


def load_default_data(data_dir: str):
    """Load assumptions/assets/deps from data/ if present, else Python defaults."""
    a_path = os.path.join(data_dir, "default_assumptions.json")
    assets_path = os.path.join(data_dir, "default_assets.csv")
    deps_path = os.path.join(data_dir, "default_dependencies.csv")

    if os.path.exists(a_path):
        with open(a_path, encoding="utf-8") as fh:
            assumptions = assumptions_from_json(fh.read())
    else:
        assumptions = defaults.default_assumptions()

    if os.path.exists(assets_path):
        with open(assets_path, encoding="utf-8") as fh:
            assets = assets_from_csv(fh.read())
    else:
        assets = defaults.default_assets()

    if os.path.exists(deps_path):
        with open(deps_path, encoding="utf-8") as fh:
            dependencies = dependencies_from_csv(fh.read())
    else:
        dependencies = defaults.default_dependencies()

    return assumptions, assets, dependencies
